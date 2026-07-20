import os
from openpyxl import load_workbook

@mcp.tool()
def add_excel_row(relative_path: str, sheet_name: str, row_values: list) -> str:
    """Append a new row to an Excel sheet.
    
    Args:
        relative_path: The relative path of the Excel file.
        sheet_name: The name of the sheet to edit.
        row_values: A list of values to append as a new row (e.g. [1, "John", 45.2, "2026-07-04"]).
    """
    safe_path = os.path.abspath(os.path.join(AI_WORKSPACE_DIR, relative_path))
    if not safe_path.startswith(os.path.abspath(AI_WORKSPACE_DIR)):
        return "Error: Access denied."
    if not os.path.exists(safe_path):
        return f"Error: File '{relative_path}' not found."
    if os.path.isdir(safe_path):
        return f"Error: The path '{relative_path}' is a directory, not a file. Please specify a file path including a filename (e.g. 'agents/JARVIS/spreadsheet.xlsx')."
        
    try:
        wb = load_workbook(safe_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
            
        ws = wb[sheet_name]
        
        # Try to cast numeric strings in row_values
        typed_values = []
        for v in row_values:
            if isinstance(v, str) and v.strip():
                try:
                    typed_values.append(int(v))
                except ValueError:
                    try:
                        typed_values.append(float(v))
                    except ValueError:
                        typed_values.append(v)
            else:
                typed_values.append(v)
                
        ws.append(typed_values)
        wb.save(safe_path)
        wb.close()
        return f"Successfully appended row to sheet '{sheet_name}' in '{relative_path}'."
    except Exception as e:
        return f"Error adding row: {str(e)}"
