import os
from openpyxl import load_workbook

@mcp.tool()
def add_excel_sheet(relative_path: str, sheet_name: str) -> str:
    """Add a new empty worksheet to an existing Excel file.
    
    Args:
        relative_path: The relative path of the Excel file.
        sheet_name: The name of the new sheet to add.
    """
    safe_path = os.path.abspath(os.path.join(AI_WORKSPACE_DIR, relative_path))
    if not safe_path.startswith(os.path.abspath(AI_WORKSPACE_DIR)):
        return "Error: Access denied."
    if not os.path.exists(safe_path):
        return f"Error: File '{relative_path}' not found."
    if os.path.isdir(safe_path):
        return f"Error: The path '{relative_path}' is a directory, not a file. Please specify a file path including a filename (e.g. 'agents/JARVIS/spreadsheet.xlsx')."
        
    try:
        wb = load_workbook(safe_path)
        if sheet_name in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' already exists."
            
        wb.create_sheet(title=sheet_name)
        wb.save(safe_path)
        wb.close()
        return f"Successfully added sheet '{sheet_name}' to '{relative_path}'."
    except Exception as e:
        return f"Error adding sheet: {str(e)}"
