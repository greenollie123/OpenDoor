import os
from openpyxl import load_workbook

@mcp.tool()
def delete_excel_sheet(relative_path: str, sheet_name: str) -> str:
    """Delete an existing sheet from an Excel file.
    
    Args:
        relative_path: The relative path of the Excel file.
        sheet_name: The name of the sheet to delete.
    """
    safe_path = os.path.abspath(os.path.join(AI_WORKSPACE_DIR, relative_path))
    if not safe_path.startswith(os.path.abspath(AI_WORKSPACE_DIR)):
        return "Error: Access denied."
    if not os.path.exists(safe_path):
        return f"Error: File '{relative_path}' not found."
    if os.path.isdir(safe_path):
        return f"Error: The path '{relative_path}' is a directory, not a file. Please specify a file path including a filename (e.g. 'agents/JARVIS/spreadsheet.xlsx')."
        
    try:
        wb = load_workbook(safe_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
            
        if len(wb.sheetnames) <= 1:
            wb.close()
            return "Error: An Excel workbook must contain at least one visible worksheet."
            
        wb.remove(wb[sheet_name])
        wb.save(safe_path)
        wb.close()
        return f"Successfully deleted sheet '{sheet_name}' from '{relative_path}'."
    except Exception as e:
        return f"Error deleting sheet: {str(e)}"
