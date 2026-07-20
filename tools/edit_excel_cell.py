import os
from openpyxl import load_workbook

@mcp.tool()
def edit_excel_cell(relative_path: str, sheet_name: str, cell: str, value: str) -> str:
    """Edit or update a single cell in an Excel sheet (e.g. cell='A1', value='100').
    
    Args:
        relative_path: The relative path of the Excel file.
        sheet_name: The name of the sheet to edit.
        cell: The cell coordinate (e.g. 'A1', 'B3', 'C10').
        value: The new value to set. Will be cast to integer or float if possible.
    """
    safe_path = os.path.abspath(os.path.join(AI_WORKSPACE_DIR, relative_path))
    if not safe_path.startswith(os.path.abspath(AI_WORKSPACE_DIR)):
        return "Error: Access denied."
    if not os.path.exists(safe_path):
        return f"Error: File '{relative_path}' not found."
    if os.path.isdir(safe_path):
        return f"Error: The path '{relative_path}' is a directory, not a file. Please specify a file path including a filename (e.g. 'agents/JARVIS/spreadsheet.xlsx')."
        
    try:
        wb = load_workbook(safe_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
            
        ws = wb[sheet_name]
        
        # Try to convert value to numeric types if appropriate
        typed_value = value
        if isinstance(value, str) and value.strip():
            # Try int
            try:
                typed_value = int(value)
            except ValueError:
                # Try float
                try:
                    typed_value = float(value)
                except ValueError:
                    # Keep as string
                    pass
        elif isinstance(value, str):
            typed_value = None  # Clear cell if empty string is passed
            
        ws[cell] = typed_value
        wb.save(safe_path)
        wb.close()
        return f"Successfully set cell '{cell}' to '{value}' on sheet '{sheet_name}' in '{relative_path}'."
    except Exception as e:
        return f"Error editing cell: {str(e)}"
