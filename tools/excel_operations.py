import os
import pandas as pd
from openpyxl import load_workbook, Workbook

# Context injected by mcp_server.py:
# - mcp: FastMCP instance
# - AI_WORKSPACE_DIR: Root directory of agent's workspace

@mcp.tool()
def create_excel_file(relative_path: str, sheet_name: str = "Sheet1", headers: list = None) -> str:
    """Create a new Excel file (.xlsx) with an optional sheet name and headers.
    
    Args:
        relative_path: The relative path where the Excel file should be created.
        sheet_name: The name of the initial sheet (default is 'Sheet1').
        headers: Optional. A list of header names to write in the first row.
    """
    safe_path = os.path.abspath(os.path.join(AI_WORKSPACE_DIR, relative_path))
    if not safe_path.startswith(os.path.abspath(AI_WORKSPACE_DIR)):
        return "Error: Access denied."
    if os.path.isdir(safe_path):
        return f"Error: The path '{relative_path}' is an existing directory. Please specify a file path including a filename (e.g. 'agents/JARVIS/spreadsheet.xlsx')."
    
    try:
        # Create directories if they do not exist
        os.makedirs(os.path.dirname(safe_path), exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if headers:
            ws.append(headers)
            
        wb.save(safe_path)
        wb.close()
        return f"Successfully created Excel file at '{relative_path}' with sheet '{sheet_name}'."
    except Exception as e:
        return f"Error creating Excel file: {str(e)}"

@mcp.tool()
def edit_excel_cell(relative_path: str, sheet_name: str, cell: str, value: str) -> str:
    """Edit or update a single cell in an Excel sheet (e.g. cell='A1', value='100').
    
    Args:
        relative_path: The relative path of the Excel file.
        sheet_name: The name of the sheet to edit.
        cell: The cell coordinate (e.g. 'A1', 'B3', 'C10').
        value: The new value to set. Will be cast to integer or float if possible.
    """
    safe_path = os.path.abspath(os.path.join(AI_WORKSPACE_DIR, relative_path))
    if not safe_path.startswith(os.path.abspath(AI_WORKSPACE_DIR)):
        return "Error: Access denied."
    if not os.path.exists(safe_path):
        return f"Error: File '{relative_path}' not found."
    if os.path.isdir(safe_path):
        return f"Error: The path '{relative_path}' is a directory, not a file. Please specify a file path including a filename (e.g. 'agents/JARVIS/spreadsheet.xlsx')."
        
    try:
        wb = load_workbook(safe_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
            
        ws = wb[sheet_name]
        
        # Try to convert value to numeric types if appropriate
        typed_value = value
        if isinstance(value, str) and value.strip():
            # Try int
            try:
                typed_value = int(value)
            except ValueError:
                # Try float
                try:
                    typed_value = float(value)
                except ValueError:
                    # Keep as string
                    pass
        elif isinstance(value, str):
            typed_value = None  # Clear cell if empty string is passed
            
        ws[cell] = typed_value
        wb.save(safe_path)
        wb.close()
        return f"Successfully set cell '{cell}' to '{value}' on sheet '{sheet_name}' in '{relative_path}'."
    except Exception as e:
        return f"Error editing cell: {str(e)}"

@mcp.tool()
def add_excel_row(relative_path: str, sheet_name: str, row_values: list) -> str:
    """Append a new row to an Excel sheet.
    
    Args:
        relative_path: The relative path of the Excel file.
        sheet_name: The name of the sheet to edit.
        row_values: A list of values to append as a new row (e.g. [1, "John", 45.2, "2026-07-04"]).
    """
    safe_path = os.path.abspath(os.path.join(AI_WORKSPACE_DIR, relative_path))
    if not safe_path.startswith(os.path.abspath(AI_WORKSPACE_DIR)):
        return "Error: Access denied."
    if not os.path.exists(safe_path):
        return f"Error: File '{relative_path}' not found."
    if os.path.isdir(safe_path):
        return f"Error: The path '{relative_path}' is a directory, not a file. Please specify a file path including a filename (e.g. 'agents/JARVIS/spreadsheet.xlsx')."
        
    try:
        wb = load_workbook(safe_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
            
        ws = wb[sheet_name]
        
        # Try to cast numeric strings in row_values
        typed_values = []
        for v in row_values:
            if isinstance(v, str) and v.strip():
                try:
                    typed_values.append(int(v))
                except ValueError:
                    try:
                        typed_values.append(float(v))
                    except ValueError:
                        typed_values.append(v)
            else:
                typed_values.append(v)
                
        ws.append(typed_values)
        wb.save(safe_path)
        wb.close()
        return f"Successfully appended row to sheet '{sheet_name}' in '{relative_path}'."
    except Exception as e:
        return f"Error adding row: {str(e)}"

@mcp.tool()
def add_excel_sheet(relative_path: str, sheet_name: str) -> str:
    """Add a new empty worksheet to an existing Excel file.
    
    Args:
        relative_path: The relative path of the Excel file.
        sheet_name: The name of the new sheet to add.
    """
    safe_path = os.path.abspath(os.path.join(AI_WORKSPACE_DIR, relative_path))
    if not safe_path.startswith(os.path.abspath(AI_WORKSPACE_DIR)):
        return "Error: Access denied."
    if not os.path.exists(safe_path):
        return f"Error: File '{relative_path}' not found."
    if os.path.isdir(safe_path):
        return f"Error: The path '{relative_path}' is a directory, not a file. Please specify a file path including a filename (e.g. 'agents/JARVIS/spreadsheet.xlsx')."
        
    try:
        wb = load_workbook(safe_path)
        if sheet_name in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' already exists."
            
        wb.create_sheet(title=sheet_name)
        wb.save(safe_path)
        wb.close()
        return f"Successfully added sheet '{sheet_name}' to '{relative_path}'."
    except Exception as e:
        return f"Error adding sheet: {str(e)}"

@mcp.tool()
def delete_excel_sheet(relative_path: str, sheet_name: str) -> str:
    """Delete an existing sheet from an Excel file.
    
    Args:
        relative_path: The relative path of the Excel file.
        sheet_name: The name of the sheet to delete.
    """
    safe_path = os.path.abspath(os.path.join(AI_WORKSPACE_DIR, relative_path))
    if not safe_path.startswith(os.path.abspath(AI_WORKSPACE_DIR)):
        return "Error: Access denied."
    if not os.path.exists(safe_path):
        return f"Error: File '{relative_path}' not found."
    if os.path.isdir(safe_path):
        return f"Error: The path '{relative_path}' is a directory, not a file. Please specify a file path including a filename (e.g. 'agents/JARVIS/spreadsheet.xlsx')."
        
    try:
        wb = load_workbook(safe_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
            
        if len(wb.sheetnames) <= 1:
            wb.close()
            return "Error: An Excel workbook must contain at least one visible worksheet."
            
        wb.remove(wb[sheet_name])
        wb.save(safe_path)
        wb.close()
        return f"Successfully deleted sheet '{sheet_name}' from '{relative_path}'."
    except Exception as e:
        return f"Error deleting sheet: {str(e)}"

@mcp.tool()
def update_excel_sheet_from_markdown(relative_path: str, sheet_name: str, markdown_table: str) -> str:
    """Overwrite or update an Excel sheet using a Markdown table string.
    This allows updating the whole sheet by editing its Markdown representation.
    
    Args:
        relative_path: The relative path of the Excel file.
        sheet_name: The name of the sheet to overwrite or create.
        markdown_table: The complete markdown table representing the sheet content (e.g. '| Col1 | Col2 |\\n|---|---|\\n| Val1 | Val2 |').
    """
    safe_path = os.path.abspath(os.path.join(AI_WORKSPACE_DIR, relative_path))
    if not safe_path.startswith(os.path.abspath(AI_WORKSPACE_DIR)):
        return "Error: Access denied."
    if os.path.isdir(safe_path):
        return f"Error: The path '{relative_path}' is an existing directory. Please specify a file path including a filename (e.g. 'agents/JARVIS/spreadsheet.xlsx')."
        
    try:
        # Parse the markdown table
        lines = [line.strip() for line in markdown_table.strip().split('\n') if line.strip()]
        if not lines:
            return "Error: Empty markdown table."
            
        cleaned_lines = []
        for line in lines:
            # Strip outer pipes
            if line.startswith('|'):
                line = line[1:]
            if line.endswith('|'):
                line = line[:-1]
            cleaned_lines.append([val.strip() for val in line.split('|')])
            
        if len(cleaned_lines) < 1:
            return "Error: Invalid markdown table format."
            
        # Check for separation line (e.g., |---|---|)
        has_separator = len(cleaned_lines) > 1 and all(all(c in '-:' for c in col) for col in cleaned_lines[1] if col)
        
        if has_separator:
            headers = [h for h in cleaned_lines[0]]
            data = cleaned_lines[2:]
        else:
            headers = [h for h in cleaned_lines[0]]
            data = cleaned_lines[1:]
            
        # Filter data rows to make sure they match header length
        num_cols = len(headers)
        fixed_data = []
        for row in data:
            if len(row) < num_cols:
                row = row + [""] * (num_cols - len(row))
            elif len(row) > num_cols:
                row = row[:num_cols]
            fixed_data.append(row)
            
        df = pd.DataFrame(fixed_data, columns=headers)
        
        # Replace empty strings with None so they show as blank cells in Excel
        df = df.replace(to_replace={"": None})
        
        # Try to convert columns to numeric types where possible
        for col in df.columns:
            try:
                converted = pd.to_numeric(df[col], errors='raise')
                df[col] = converted
            except (ValueError, TypeError):
                pass
                
        # Write to Excel file
        if os.path.exists(safe_path):
            wb = load_workbook(safe_path)
            if sheet_name in wb.sheetnames:
                if len(wb.sheetnames) > 1:
                    wb.remove(wb[sheet_name])
                    wb.save(safe_path)
                    wb.close()
                    with pd.ExcelWriter(safe_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    wb.close()
                    with pd.ExcelWriter(safe_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                wb.close()
                with pd.ExcelWriter(safe_path, mode='a', engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            os.makedirs(os.path.dirname(safe_path), exist_ok=True)
            with pd.ExcelWriter(safe_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
        return f"Successfully updated sheet '{sheet_name}' in '{relative_path}' using the markdown table."
    except Exception as e:
        return f"Error updating sheet from markdown: {str(e)}"

@mcp.tool()
def read_excel_file(relative_path: str, sheet_name: str = None) -> str:
    """Reads the contents of an Excel file (.xlsx or .xls) and returns it as a formatted Markdown string.
    
    Args:
        relative_path: The relative path of the Excel file inside the workspace.
        sheet_name: Optional. The name of the specific sheet to read. If not specified, reads all sheets.
    """
    safe_path = os.path.abspath(os.path.join(AI_WORKSPACE_DIR, relative_path))
    if not safe_path.startswith(os.path.abspath(AI_WORKSPACE_DIR)):
        return "Error: Access denied."
    if not os.path.exists(safe_path):
        return f"Error: File '{relative_path}' not found."
    if os.path.isdir(safe_path):
        return f"Error: The path '{relative_path}' is a directory, not a file. Please specify a file path including a filename (e.g. 'agents/JARVIS/spreadsheet.xlsx')."
    
    try:
        # Load file using pandas ExcelFile in a context manager to auto-close handles
        with pd.ExcelFile(safe_path) as xl:
            sheets = xl.sheet_names
            
            output = []
            if sheet_name:
                if sheet_name not in sheets:
                    return f"Error: Sheet '{sheet_name}' not found. Available sheets: {', '.join(sheets)}"
                sheets_to_read = [sheet_name]
            else:
                sheets_to_read = sheets
                
            for sheet in sheets_to_read:
                df = pd.read_excel(xl, sheet_name=sheet)
                df = df.fillna("")
                output.append(f"### Sheet: {sheet}")
                if df.empty:
                    output.append("*(This sheet is empty)*")
                else:
                    total_rows = len(df)
                    if total_rows > 100:
                        df_subset = df.head(100)
                        output.append(df_subset.to_markdown(index=False))
                        output.append(f"\n*(Truncated: showing first 100 of {total_rows} rows. Use sheet filtering if needed.)*")
                    else:
                        output.append(df.to_markdown(index=False))
                output.append("")
                
            return "\n".join(output)
    except Exception as e:
        return f"Error reading Excel file: {str(e)}"
