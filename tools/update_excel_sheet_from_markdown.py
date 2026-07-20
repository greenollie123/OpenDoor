import os
import pandas as pd
from openpyxl import load_workbook

@mcp.tool()
def update_excel_sheet_from_markdown(relative_path: str, sheet_name: str, markdown_table: str) -> str:
    """Overwrite or update an Excel sheet using a Markdown table string.
    This allows updating the whole sheet by editing its Markdown representation.
    
    Args:
        relative_path: The relative path of the Excel file.
        sheet_name: The name of the sheet to overwrite or create.
        markdown_table: The complete markdown table representing the sheet content (e.g. '| Col1 | Col2 |\\n|---|---|\\n| Val1 | Val2 |').
    """
    safe_path = os.path.abspath(os.path.join(AI_WORKSPACE_DIR, relative_path))
    if not safe_path.startswith(os.path.abspath(AI_WORKSPACE_DIR)):
        return "Error: Access denied."
    if os.path.isdir(safe_path):
        return f"Error: The path '{relative_path}' is an existing directory. Please specify a file path including a filename (e.g. 'agents/JARVIS/spreadsheet.xlsx')."
        
    try:
        # Parse the markdown table
        lines = [line.strip() for line in markdown_table.strip().split('\n') if line.strip()]
        if not lines:
            return "Error: Empty markdown table."
            
        cleaned_lines = []
        for line in lines:
            # Strip outer pipes
            if line.startswith('|'):
                line = line[1:]
            if line.endswith('|'):
                line = line[:-1]
            cleaned_lines.append([val.strip() for val in line.split('|')])
            
        if len(cleaned_lines) < 1:
            return "Error: Invalid markdown table format."
            
        # Check for separation line (e.g., |---|---|)
        has_separator = len(cleaned_lines) > 1 and all(all(c in '-:' for c in col) for col in cleaned_lines[1] if col)
        
        if has_separator:
            headers = [h for h in cleaned_lines[0]]
            data = cleaned_lines[2:]
        else:
            headers = [h for h in cleaned_lines[0]]
            data = cleaned_lines[1:]
            
        # Filter data rows to make sure they match header length
        num_cols = len(headers)
        fixed_data = []
        for row in data:
            if len(row) < num_cols:
                row = row + [""] * (num_cols - len(row))
            elif len(row) > num_cols:
                row = row[:num_cols]
            fixed_data.append(row)
            
        df = pd.DataFrame(fixed_data, columns=headers)
        
        # Replace empty strings with None so they show as blank cells in Excel
        df = df.replace(to_replace={"": None})
        
        # Try to convert columns to numeric types where possible
        for col in df.columns:
            try:
                converted = pd.to_numeric(df[col], errors='raise')
                df[col] = converted
            except (ValueError, TypeError):
                pass
                
        # Write to Excel file
        if os.path.exists(safe_path):
            wb = load_workbook(safe_path)
            if sheet_name in wb.sheetnames:
                if len(wb.sheetnames) > 1:
                    wb.remove(wb[sheet_name])
                    wb.save(safe_path)
                    wb.close()
                    with pd.ExcelWriter(safe_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    wb.close()
                    with pd.ExcelWriter(safe_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                wb.close()
                with pd.ExcelWriter(safe_path, mode='a', engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            os.makedirs(os.path.dirname(safe_path), exist_ok=True)
            with pd.ExcelWriter(safe_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
        return f"Successfully updated sheet '{sheet_name}' in '{relative_path}' using the markdown table."
    except Exception as e:
        return f"Error updating sheet from markdown: {str(e)}"
